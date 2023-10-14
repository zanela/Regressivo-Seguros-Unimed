import openpyxl

global workbook
workbook = openpyxl.load_workbook("C:/Automação/Portal_Cliente_PF/Data/Massa_de_dados.xlsx")
global baseWorkbook
baseWorkbook = openpyxl.load_workbook("C:/Automação/Portal_Cliente_PF/Data/base.xlsx")

global sheet
sheet = workbook["Dados de acesso"]
global baseSheet
baseSheet = baseWorkbook["Sheet1"]


def get_username_and_password(segment):
    username = globals()
    password = globals()
    rows = sheet.max_row
    columns = sheet.max_column

    for i in range(1, rows + 1):
        for j in range(1, columns + 1):
            cell = sheet.cell(i, j)
            if cell.value == segment:
                usr = sheet.cell(i, 2)
                pwd = sheet.cell(i, 3)
                username = str(usr.value)
                password = str(pwd.value)
    return [username, password]


def get_url():
    global url
    rows = baseSheet.max_row
    columns = baseSheet.max_column

    for i in range(1, rows + 1):
        for j in range(1, columns + 1):
            cell = baseSheet.cell(i, j)
            if cell.value == 'URL':
                cellurl = baseSheet.cell(i, 2)
                url = str(cellurl.value)
    return url


def get_evidences_path():
    global evidencespath
    rows = baseSheet.max_row
    columns = baseSheet.max_column

    for i in range(1, rows + 1):
        for j in range(1, columns + 1):
            cell = baseSheet.cell(i, j)
            if cell.value == 'evidencesPath':
                cellevidencespath = baseSheet.cell(i, 2)
                evidencespath = str(cellevidencespath.value)
    return evidencespath



